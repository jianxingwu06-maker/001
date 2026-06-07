from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .models import HUMAN_CONFIRM, ListingVersion

HEADER_ALIASES = {
    "商品标题": ["商品标题", "标题", "产品标题", "Product Title"],
    "商品描述": ["商品描述", "描述", "产品描述", "Description"],
    "SKU": ["SKU", "商家编码", "商品编码"],
    "规格名": ["规格名", "规格名称", "Variation Name"],
    "规格值": ["规格值", "规格", "Variation Value"],
    "颜色": ["颜色", "Color"],
    "尺寸": ["尺寸", "尺码", "Size"],
    "重量": ["重量", "Weight"],
    "体积": ["体积", "包装尺寸", "Package Size"],
    "类目 ID": ["类目 ID", "类目ID", "Category ID", "类目"],
    "属性": ["属性", "Attributes"],
    "主图": ["主图", "Main Image", "图片1"],
    "附图": ["附图", "副图", "Gallery Images", "图片2"],
    "详情图": ["详情图", "Detail Images"],
    "库存": ["库存", "Stock"],
    "售价": ["售价", "价格", "Price"],
    "物流信息": ["物流信息", "物流", "Logistics"],
}


class DxmExcelWriter:
    def __init__(self, template_path: Path, output_path: Path):
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)

    def write(self, versions: list[ListingVersion]) -> dict[str, Any]:
        wb = load_workbook(self.template_path)
        ws = wb.active
        header_row, headers = self._find_header(ws)
        mapping = self._map_headers(headers)
        missing_required = []
        row = self._first_empty_row(ws, header_row, headers)
        for version in versions:
            self._copy_row_style(ws, row - 1 if row > header_row + 1 else header_row, row)
            values = self._version_values(version)
            for canonical, value in values.items():
                col = mapping.get(canonical)
                if col:
                    ws.cell(row=row, column=col).value = value
                elif canonical in ["商品标题", "商品描述", "SKU", "主图", "售价", "库存"]:
                    missing_required.append(canonical)
            row += 1
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        return {"output": str(self.output_path), "missing_template_columns": sorted(set(missing_required))}

    def _find_header(self, ws) -> tuple[int, dict[int, str]]:
        for row in range(1, min(ws.max_row, 30) + 1):
            headers = {cell.column: str(cell.value).strip() for cell in ws[row] if cell.value}
            normalized = " ".join(headers.values())
            if any(alias in normalized for alias in ["商品标题", "SKU", "主图", "售价", "库存", "Product Title"]):
                return row, headers
        raise ValueError("未找到店小秘表头行，请确认模板包含商品标题/SKU/主图等字段")

    def _map_headers(self, headers: dict[int, str]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        lowered = {col: name.lower().replace(" ", "") for col, name in headers.items()}
        for canonical, aliases in HEADER_ALIASES.items():
            for col, header in lowered.items():
                if any(alias.lower().replace(" ", "") in header for alias in aliases):
                    mapping[canonical] = col
                    break
        return mapping

    def _first_empty_row(self, ws, header_row: int, headers: dict[int, str]) -> int:
        key_col = next(iter(headers.keys()))
        row = header_row + 1
        while ws.cell(row=row, column=key_col).value not in (None, ""):
            row += 1
        return row

    def _copy_row_style(self, ws, source_row: int, target_row: int) -> None:
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)

    def _version_values(self, v: ListingVersion) -> dict[str, Any]:
        return {
            "商品标题": v.title,
            "商品描述": v.description,
            "SKU": v.sku,
            "规格名": v.spec_name,
            "规格值": v.spec_value,
            "颜色": v.color,
            "尺寸": v.size,
            "重量": v.weight,
            "体积": v.package_size,
            "类目 ID": v.category_id,
            "属性": pd.Series(v.attributes, dtype="object").to_json(force_ascii=False),
            "主图": v.main_image,
            "附图": ";".join(v.gallery_images),
            "详情图": ";".join(v.detail_images),
            "库存": v.stock,
            "售价": v.price,
            "物流信息": v.logistics,
        }
